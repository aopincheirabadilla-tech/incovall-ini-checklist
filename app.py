# -*- coding: utf-8 -*-
from flask import Flask, render_template, request, redirect, url_for, send_file, flash
from flask_login import (LoginManager, UserMixin, login_user, logout_user,
                         login_required, current_user)
from werkzeug.security import generate_password_hash, check_password_hash
from functools import wraps
import sqlite3
import os
import io
from datetime import datetime
from io import BytesIO

from PIL import Image as PILImage

from reportlab.lib.pagesizes import A4
from reportlab.lib import colors
from reportlab.lib.units import cm
from reportlab.platypus import SimpleDocTemplate, Table, TableStyle, Paragraph, Spacer, Image
from reportlab.lib.styles import getSampleStyleSheet, ParagraphStyle
from reportlab.lib.enums import TA_CENTER, TA_LEFT
from reportlab.graphics.shapes import Drawing, Rect, String as RLString

app = Flask(__name__)
app.secret_key = os.environ.get('SECRET_KEY', 'incovall-ini-2024-auth')
app.config['MAX_CONTENT_LENGTH'] = 80 * 1024 * 1024  # 80 MB total por request

# ─── Flask-Login ─────────────────────────────────────────────────────────────

login_manager = LoginManager(app)
login_manager.login_view = 'login'
login_manager.login_message = 'Debes iniciar sesión para acceder.'
login_manager.login_message_category = 'warning'

ROLES  = ['admin', 'supervisor', 'inspector', 'lectura']
FAENAS = ['PLANTA MAGNETITA', 'CNN', 'PTT']
ALLOWED_MIME = {'image/jpeg', 'image/png', 'image/webp'}

# ─── Normalización de faenas ──────────────────────────────────────────────────

_FAENA_VARIANTES = {
    'PLANTA MAGNETITA': {
        'magnetita', 'planta magnetita', 'pm', 'p.m.', 'planta mag',
    },
    'CNN': {
        'cnn', 'cerro negro', 'cerro negro norte',
    },
    'PTT': {
        'ptt', 'ppt', 'puerto', 'puerto punta', 'totoralillo', 'totoralilo',
        'totoralollo', 'puerto punta totoralillo', 'punta totoralillo',
        'puerto punta totoralollo', 'punta totoralollo', 'puerto totoralollo',
        'puerto punta totoralilo', 'punta totoralilo',
    },
}
# Índice invertido: variante_en_minúsculas → nombre_oficial
_FAENA_INDEX: dict = {}
for _oficial, _vars in _FAENA_VARIANTES.items():
    _FAENA_INDEX[_oficial.lower()] = _oficial   # el propio nombre oficial
    for _v in _vars:
        _FAENA_INDEX[_v.lower()] = _oficial


def normalizar_faena(texto: str) -> str:
    """Devuelve el nombre oficial de la faena o el texto en mayúsculas si no se reconoce."""
    if not texto:
        return ''
    return _FAENA_INDEX.get(texto.strip().lower(), texto.strip().upper())
ALLOWED_EXT  = {'jpg', 'jpeg', 'png', 'webp'}
MAX_FOTO_BYTES = 5 * 1024 * 1024  # 5 MB por foto


class User(UserMixin):
    def __init__(self, row):
        self.id              = row['id']
        self.username        = row['username']
        self.nombre_completo = row['nombre_completo']
        self.email           = row['email']
        self.password_hash   = row['password_hash']
        self.rol             = row['rol']
        self.faena           = row['faena']
        self.activo          = row['activo']

    def get_id(self):
        return str(self.id)

    @property
    def is_active(self):
        return bool(self.activo)

    def is_admin(self):
        return self.rol == 'admin'

    def is_supervisor(self):
        return self.rol in ('admin', 'supervisor')

    def is_lectura(self):
        return self.rol == 'lectura'

    def puede_ver(self):
        """Puede acceder a vistas de solo lectura (dashboard, historial, acciones, requerimientos)."""
        return True  # todos los roles autenticados pueden ver

    def puede_escribir(self):
        """Puede crear/editar/modificar registros."""
        return self.rol != 'lectura'


@login_manager.user_loader
def load_user(user_id):
    conn = get_db()
    row = conn.execute('SELECT * FROM usuarios WHERE id=?', (user_id,)).fetchone()
    conn.close()
    return User(row) if row else None


def admin_required(f):
    @wraps(f)
    def decorated(*args, **kwargs):
        if not current_user.is_authenticated or not current_user.is_admin():
            flash('Acceso restringido a administradores.', 'danger')
            return redirect(url_for('index'))
        return f(*args, **kwargs)
    return decorated


def escritura_required(f):
    """Bloquea el acceso a usuarios con rol 'lectura'."""
    @wraps(f)
    def decorated(*args, **kwargs):
        if current_user.is_authenticated and not current_user.puede_escribir():
            flash('Acceso restringido: su rol es de solo lectura.', 'warning')
            return redirect(url_for('dashboard'))
        return f(*args, **kwargs)
    return decorated


def _puede_ver_inspeccion(insp):
    """True si el usuario actual tiene permiso para ver esta inspección."""
    if current_user.is_admin():
        return True
    if current_user.rol in ('supervisor', 'lectura'):
        return insp['faena'] == current_user.faena or current_user.faena is None
    return insp['usuario_id'] == current_user.id


# ─── Config ──────────────────────────────────────────────────────────────────

BASE_DIR   = os.path.dirname(__file__)
DATABASE   = os.environ.get('DATABASE_PATH', os.path.join(BASE_DIR, 'inspecciones.db'))
STATIC_DIR = os.path.join(BASE_DIR, 'static')
FOTOS_DIR  = os.environ.get('FOTOS_PATH', os.path.join(STATIC_DIR, 'fotos'))

SECCIONES = {
    'SISTEMA SANITARIO': ['WC', 'Lavamanos', 'Duchas', 'Agua Fría', 'Agua Caliente', 'Termo', 'Filtraciones'],
    'SISTEMA ELÉCTRICO': ['Equipos de Iluminación', 'Enchufes', 'Interruptores'],
    'INFRAESTRUCTURA GENERAL': ['Muros', 'Pisos', 'Cielos'],
    'SISTEMA DE ALCANTARILLADO': ['Cámaras', 'Tuberías'],
}
FLAT_ITEMS = [(s, item) for s, items in SECCIONES.items() for item in items]

@app.template_filter('seccion_icon')
def seccion_icon(value):
    iconos = ['droplet-half', 'lightning-charge', 'building', 'pipe']
    try:
        return iconos[int(value) - 1]
    except Exception:
        return 'check-circle'


# ─── Database ────────────────────────────────────────────────────────────────

def get_db():
    conn = sqlite3.connect(DATABASE)
    conn.row_factory = sqlite3.Row
    return conn


def init_db():
    os.makedirs(FOTOS_DIR, exist_ok=True)
    conn = get_db()
    c = conn.cursor()

    c.execute('''
        CREATE TABLE IF NOT EXISTS usuarios (
            id               INTEGER PRIMARY KEY AUTOINCREMENT,
            username         TEXT UNIQUE NOT NULL,
            nombre_completo  TEXT NOT NULL,
            email            TEXT,
            password_hash    TEXT NOT NULL,
            rol              TEXT NOT NULL DEFAULT 'inspector',
            faena            TEXT,
            activo           INTEGER NOT NULL DEFAULT 1,
            created_at       TIMESTAMP DEFAULT CURRENT_TIMESTAMP
        )
    ''')

    c.execute('''
        CREATE TABLE IF NOT EXISTS inspecciones (
            id               INTEGER PRIMARY KEY AUTOINCREMENT,
            correlativo      TEXT UNIQUE,
            valle            TEXT,
            faena            TEXT,
            fecha            TEXT,
            nombre_recinto   TEXT,
            nombre_inspector TEXT,
            cliente          TEXT,
            observaciones    TEXT,
            supervisor       TEXT,
            usuario_id       INTEGER,
            created_at       TIMESTAMP DEFAULT CURRENT_TIMESTAMP,
            FOREIGN KEY (usuario_id) REFERENCES usuarios(id)
        )
    ''')

    # Migración: usuario_id en inspecciones viejas
    cols = [r[1] for r in c.execute("PRAGMA table_info(inspecciones)").fetchall()]
    if 'usuario_id' not in cols:
        c.execute("ALTER TABLE inspecciones ADD COLUMN usuario_id INTEGER")

    c.execute('''
        CREATE TABLE IF NOT EXISTS items_checklist (
            id            INTEGER PRIMARY KEY AUTOINCREMENT,
            inspeccion_id INTEGER,
            seccion       TEXT,
            item          TEXT,
            valor         TEXT,
            comentario    TEXT,
            FOREIGN KEY (inspeccion_id) REFERENCES inspecciones(id)
        )
    ''')

    c.execute('''
        CREATE TABLE IF NOT EXISTS fotos_items (
            id            INTEGER PRIMARY KEY AUTOINCREMENT,
            inspeccion_id INTEGER NOT NULL,
            item_idx      INTEGER NOT NULL,
            filename      TEXT NOT NULL,
            created_at    TIMESTAMP DEFAULT CURRENT_TIMESTAMP,
            FOREIGN KEY (inspeccion_id) REFERENCES inspecciones(id)
        )
    ''')

    c.execute('''
        CREATE TABLE IF NOT EXISTS acciones_correctivas (
            id              INTEGER PRIMARY KEY AUTOINCREMENT,
            inspeccion_id   INTEGER NOT NULL,
            item_idx        INTEGER NOT NULL,
            recinto         TEXT NOT NULL,
            faena           TEXT,
            item            TEXT NOT NULL,
            descripcion     TEXT,
            foto_filename   TEXT,
            responsable_id  INTEGER,
            fecha_limite    TEXT,
            prioridad       TEXT NOT NULL DEFAULT 'Media',
            estado          TEXT NOT NULL DEFAULT 'Pendiente',
            observaciones   TEXT,
            fecha_cierre    TEXT,
            created_at      TIMESTAMP DEFAULT CURRENT_TIMESTAMP,
            FOREIGN KEY (inspeccion_id) REFERENCES inspecciones(id),
            FOREIGN KEY (responsable_id) REFERENCES usuarios(id)
        )
    ''')

    c.execute('''
        CREATE TABLE IF NOT EXISTS seguimiento_acciones (
            id          INTEGER PRIMARY KEY AUTOINCREMENT,
            accion_id   INTEGER NOT NULL,
            usuario_id  INTEGER NOT NULL,
            observacion TEXT NOT NULL,
            estado_nuevo TEXT,
            created_at  TIMESTAMP DEFAULT CURRENT_TIMESTAMP,
            FOREIGN KEY (accion_id)  REFERENCES acciones_correctivas(id),
            FOREIGN KEY (usuario_id) REFERENCES usuarios(id)
        )
    ''')

    c.execute('''
        CREATE TABLE IF NOT EXISTS requerimientos (
            id                  INTEGER PRIMARY KEY AUTOINCREMENT,
            item_nombre         TEXT NOT NULL,
            cantidad            INTEGER NOT NULL DEFAULT 1,
            faena               TEXT,
            recinto             TEXT,
            fecha_ultimo_reporte TEXT,
            estado              TEXT NOT NULL DEFAULT 'Pendiente',
            created_at          TIMESTAMP DEFAULT CURRENT_TIMESTAMP
        )
    ''')

    # Migración: normalizar faenas existentes en todas las tablas
    for tabla in ('inspecciones', 'acciones_correctivas', 'requerimientos'):
        rows = c.execute(f'SELECT id, faena FROM {tabla}').fetchall()
        for row in rows:
            norm = normalizar_faena(row[1] or '')
            if norm != (row[1] or ''):
                c.execute(f'UPDATE {tabla} SET faena=? WHERE id=?', (norm, row[0]))

    # Migración: normalizar recinto a Title Case en todas las tablas
    for tabla, campo in (('inspecciones', 'nombre_recinto'),
                         ('acciones_correctivas', 'recinto'),
                         ('requerimientos', 'recinto')):
        rows = c.execute(f'SELECT id, {campo} FROM {tabla}').fetchall()
        for row in rows:
            val = row[1] or ''
            titled = val.title()
            if titled != val:
                c.execute(f'UPDATE {tabla} SET {campo}=? WHERE id=?', (titled, row[0]))

    # Admin inicial
    if not c.execute("SELECT id FROM usuarios WHERE username='admin'").fetchone():
        c.execute('''
            INSERT INTO usuarios (username, nombre_completo, email, password_hash, rol, faena, activo)
            VALUES (?,?,?,?,?,?,?)
        ''', ('admin', 'Administrador INCOVALL', 'admin@incovall.cl',
              generate_password_hash('incovall2026'), 'admin', None, 1))

    conn.commit()
    conn.close()


def generate_correlativo():
    conn = get_db()
    year = datetime.now().year
    row = conn.execute(
        "SELECT COUNT(*) as n FROM inspecciones WHERE correlativo LIKE ?",
        (f'INI-{year}-%',)
    ).fetchone()
    conn.close()
    return f'INI-{year}-{row["n"] + 1:04d}'


def _save_foto(file_storage, correlativo, item_idx):
    """Valida, redimensiona y guarda una foto. Devuelve el filename o None."""
    if not file_storage or not file_storage.filename:
        return None

    ext = file_storage.filename.rsplit('.', 1)[-1].lower() if '.' in file_storage.filename else ''
    mime = file_storage.content_type or ''

    if ext not in ALLOWED_EXT and mime not in ALLOWED_MIME:
        return None

    data = file_storage.read()
    if len(data) > MAX_FOTO_BYTES:
        return None

    try:
        img = PILImage.open(BytesIO(data))
        img = img.convert('RGB')
        img.thumbnail((1400, 1050), PILImage.LANCZOS)

        safe_corr = correlativo.replace('/', '_').replace('\\', '_')
        filename  = f"{safe_corr}_{item_idx}.jpg"
        filepath  = os.path.join(FOTOS_DIR, filename)
        img.save(filepath, 'JPEG', quality=82, optimize=True)
        return filename
    except Exception:
        return None


# ─── Dashboard helpers ───────────────────────────────────────────────────────

def _build_where(faena_filter, fecha_desde, fecha_hasta, alias='i'):
    """Devuelve (cláusula WHERE str, lista de parámetros) para filtros del dashboard."""
    clauses, params = [], []
    if faena_filter:
        clauses.append(f"{alias}.faena = ?"); params.append(faena_filter)
    if fecha_desde:
        clauses.append(f"{alias}.fecha >= ?"); params.append(fecha_desde)
    if fecha_hasta:
        clauses.append(f"{alias}.fecha <= ?"); params.append(fecha_hasta)
    return ("WHERE " + " AND ".join(clauses)) if clauses else "", params


def _dashboard_data(faena_filter, fecha_desde, fecha_hasta):
    """Devuelve dict con todas las estadísticas para el dashboard."""
    from datetime import timedelta
    conn = get_db()

    wh,  wp  = _build_where(faena_filter, fecha_desde, fecha_hasta, alias='i')
    # WHERE para JOIN con items (agrega c.valor='NO' al final)
    no_clauses = []
    no_params  = []
    if faena_filter: no_clauses.append("i.faena = ?");  no_params.append(faena_filter)
    if fecha_desde:  no_clauses.append("i.fecha >= ?"); no_params.append(fecha_desde)
    if fecha_hasta:  no_clauses.append("i.fecha <= ?"); no_params.append(fecha_hasta)
    no_clauses.append("c.valor = 'NO'")
    wh_no = "WHERE " + " AND ".join(no_clauses)

    # ── 1. Tarjetas ───────────────────────────────────────────────────────────
    total_inspecciones = conn.execute(
        f"SELECT COUNT(*) FROM inspecciones i {wh}", wp).fetchone()[0]
    total_no = conn.execute(
        f"SELECT COUNT(*) FROM items_checklist c "
        f"JOIN inspecciones i ON c.inspeccion_id=i.id {wh_no}", no_params).fetchone()[0]
    recintos_count = conn.execute(
        f"SELECT COUNT(DISTINCT nombre_recinto) FROM inspecciones i {wh}", wp).fetchone()[0]
    inspectores_activos = conn.execute(
        f"SELECT COUNT(DISTINCT nombre_inspector) FROM inspecciones i {wh}", wp).fetchone()[0]

    # ── 2. Top 10 ítems con más NO ────────────────────────────────────────────
    top_items = conn.execute(
        f"SELECT c.item, COUNT(*) AS cnt FROM items_checklist c "
        f"JOIN inspecciones i ON c.inspeccion_id=i.id {wh_no} "
        f"GROUP BY c.item ORDER BY cnt DESC LIMIT 10", no_params).fetchall()

    # ── 3. Distribución SI / NO / N/A ─────────────────────────────────────────
    # Reutilizar wp sin el filtro de valor
    dist_rows = conn.execute(
        f"SELECT c.valor, COUNT(*) AS cnt FROM items_checklist c "
        f"JOIN inspecciones i ON c.inspeccion_id=i.id {wh} "
        f"GROUP BY c.valor", wp).fetchall()
    dist = {'SI': 0, 'NO': 0, 'N/A': 0}
    for r in dist_rows:
        if r['valor'] in dist:
            dist[r['valor']] = r['cnt']

    # ── 4. Inspecciones por semana (últimas 8 semanas) ────────────────────────
    ref_date = datetime.strptime(fecha_hasta, '%Y-%m-%d') if fecha_hasta else datetime.now()
    semanas_labels, semanas_vals = [], []
    for w in range(7, -1, -1):
        week_end   = ref_date - timedelta(weeks=w)
        week_start = week_end - timedelta(days=6)
        w_clauses = ["fecha BETWEEN ? AND ?"]
        w_params  = [week_start.strftime('%Y-%m-%d'), week_end.strftime('%Y-%m-%d')]
        if faena_filter:
            w_clauses.append("faena = ?"); w_params.append(faena_filter)
        cnt = conn.execute(
            "SELECT COUNT(*) FROM inspecciones WHERE " + " AND ".join(w_clauses),
            w_params).fetchone()[0]
        semanas_labels.append(f"{week_start.strftime('%d/%m')}–{week_end.strftime('%d/%m')}")
        semanas_vals.append(cnt)

    # ── 5. Recintos críticos ──────────────────────────────────────────────────
    recintos_criticos = conn.execute(
        f"SELECT i.nombre_recinto, i.faena, COUNT(*) AS total_no "
        f"FROM items_checklist c JOIN inspecciones i ON c.inspeccion_id=i.id "
        f"{wh_no} GROUP BY i.nombre_recinto, i.faena "
        f"ORDER BY total_no DESC LIMIT 15", no_params).fetchall()

    # ── 6. Inspecciones por faena ─────────────────────────────────────────────
    por_faena = conn.execute(
        f"SELECT faena, COUNT(*) AS cnt FROM inspecciones i {wh} "
        f"GROUP BY faena ORDER BY cnt DESC", wp).fetchall()

    # ── 7. Acciones correctivas ───────────────────────────────────────────────
    ac_wh, ac_wp = [], []
    if faena_filter: ac_wh.append("ac.faena=?"); ac_wp.append(faena_filter)

    ac_base = ("WHERE " + " AND ".join(ac_wh)) if ac_wh else ""
    acciones_pendientes = conn.execute(
        f"SELECT COUNT(*) FROM acciones_correctivas ac {ac_base} "
        f"{'AND' if ac_base else 'WHERE'} ac.estado != 'Completada'",
        ac_wp
    ).fetchone()[0]

    today_str = datetime.now().strftime('%Y-%m-%d')
    ac_venc_wh = list(ac_wh) + ["ac.estado != 'Completada'", "ac.fecha_limite < ?"]
    ac_venc_wp = list(ac_wp) + [today_str]
    acciones_vencidas = conn.execute(
        f"SELECT ac.*, u.nombre_completo AS responsable_nombre "
        f"FROM acciones_correctivas ac "
        f"LEFT JOIN usuarios u ON ac.responsable_id = u.id "
        f"WHERE {' AND '.join(ac_venc_wh)} "
        f"ORDER BY ac.fecha_limite ASC LIMIT 10",
        ac_venc_wp
    ).fetchall()

    conn.close()
    return {
        'total_inspecciones':  total_inspecciones,
        'total_no':            total_no,
        'recintos_count':      recintos_count,
        'inspectores_activos': inspectores_activos,
        'top_items':           [(r['item'], r['cnt']) for r in top_items],
        'dist':                dist,
        'semanas_labels':      semanas_labels,
        'semanas_vals':        semanas_vals,
        'recintos_criticos':   recintos_criticos,
        'por_faena':           [(r['faena'], r['cnt']) for r in por_faena],
        'acciones_pendientes': acciones_pendientes,
        'acciones_vencidas':   acciones_vencidas,
    }


# ─── Auth Routes ─────────────────────────────────────────────────────────────

@app.route('/login', methods=['GET', 'POST'])
def login():
    if current_user.is_authenticated:
        return redirect(url_for('index'))
    if request.method == 'POST':
        username = request.form.get('username', '').strip()
        password = request.form.get('password', '')
        conn = get_db()
        row = conn.execute('SELECT * FROM usuarios WHERE username=?', (username,)).fetchone()
        conn.close()
        if row and check_password_hash(row['password_hash'], password):
            user = User(row)
            if not user.activo:
                flash('Tu cuenta está desactivada. Contacta al administrador.', 'danger')
                return render_template('login.html')
            login_user(user, remember=True)
            return redirect(request.args.get('next') or url_for('index'))
        flash('Usuario o contraseña incorrectos.', 'danger')
    return render_template('login.html')


@app.route('/logout')
@login_required
def logout():
    logout_user()
    flash('Sesión cerrada correctamente.', 'success')
    return redirect(url_for('login'))


@app.route('/perfil', methods=['GET', 'POST'])
@login_required
def perfil():
    if request.method == 'POST':
        actual      = request.form.get('password_actual', '')
        nueva       = request.form.get('password_nueva', '')
        confirmacion= request.form.get('password_confirm', '')

        conn = get_db()
        row  = conn.execute('SELECT password_hash FROM usuarios WHERE id=?',
                            (current_user.id,)).fetchone()
        conn.close()

        if not check_password_hash(row['password_hash'], actual):
            flash('La contraseña actual es incorrecta.', 'danger')
        elif len(nueva) < 6:
            flash('La nueva contraseña debe tener al menos 6 caracteres.', 'danger')
        elif nueva != confirmacion:
            flash('La confirmación no coincide con la nueva contraseña.', 'danger')
        else:
            conn = get_db()
            conn.execute('UPDATE usuarios SET password_hash=? WHERE id=?',
                         (generate_password_hash(nueva), current_user.id))
            conn.commit()
            conn.close()
            flash('Contraseña actualizada correctamente.', 'success')

    return render_template('perfil.html')


@app.route('/admin/usuarios/<int:uid>/reset_password', methods=['POST'])
@login_required
@admin_required
def admin_reset_password(uid):
    DEFAULT_PASS = 'Incovall2026*'
    conn = get_db()
    u = conn.execute('SELECT username FROM usuarios WHERE id=?', (uid,)).fetchone()
    if u is None:
        conn.close()
        flash('Usuario no encontrado.', 'danger')
        return redirect(url_for('admin_usuarios'))
    conn.execute('UPDATE usuarios SET password_hash=? WHERE id=?',
                 (generate_password_hash(DEFAULT_PASS), uid))
    conn.commit()
    conn.close()
    flash(f'Contraseña de "{u["username"]}" reseteada a: {DEFAULT_PASS}', 'success')
    return redirect(url_for('admin_usuarios'))


# ─── Main Routes ─────────────────────────────────────────────────────────────

@app.route('/')
@login_required
@escritura_required
def index():
    correlativo = generate_correlativo()
    today = datetime.now().strftime('%Y-%m-%d')
    return render_template('index.html',
                           secciones=SECCIONES,
                           flat_items=FLAT_ITEMS,
                           correlativo=correlativo,
                           today=today)


@app.route('/guardar', methods=['POST'])
@login_required
@escritura_required
def guardar():
    data = request.form
    conn = get_db()
    try:
        conn.execute('''
            INSERT INTO inspecciones
                (correlativo, valle, faena, fecha, nombre_recinto,
                 nombre_inspector, cliente, observaciones, supervisor, usuario_id)
            VALUES (?,?,?,?,?,?,?,?,?,?)
        ''', (
            data.get('correlativo'),
            data.get('valle'),
            normalizar_faena(data.get('faena', '')),
            data.get('fecha'),
            (data.get('nombre_recinto') or '').title(),
            data.get('nombre_inspector'),
            data.get('cliente'),
            data.get('observaciones'),
            data.get('supervisor'),
            current_user.id,
        ))
        inspeccion_id = conn.execute(
            "SELECT id FROM inspecciones WHERE correlativo=?",
            (data.get('correlativo'),)
        ).fetchone()['id']

        correlativo = data.get('correlativo', 'UNK')

        for i, (seccion, item) in enumerate(FLAT_ITEMS):
            valor      = data.get(f'valor_{i}', 'N/A')
            comentario = data.get(f'comentario_{i}', '')
            conn.execute('''
                INSERT INTO items_checklist (inspeccion_id, seccion, item, valor, comentario)
                VALUES (?,?,?,?,?)
            ''', (inspeccion_id, seccion, item, valor, comentario))

            # Guardar foto y crear acción correctiva si el valor es NO
            if valor == 'NO':
                foto_file = request.files.get(f'foto_{i}')
                foto_filename = _save_foto(foto_file, correlativo, i)
                if foto_filename:
                    conn.execute(
                        'INSERT INTO fotos_items (inspeccion_id, item_idx, filename) VALUES (?,?,?)',
                        (inspeccion_id, i, foto_filename)
                    )
                # Acción correctiva automática
                from datetime import timedelta
                fecha_insp  = data.get('fecha') or datetime.now().strftime('%Y-%m-%d')
                fecha_limite = (datetime.strptime(fecha_insp, '%Y-%m-%d')
                                + timedelta(days=7)).strftime('%Y-%m-%d')
                conn.execute('''
                    INSERT INTO acciones_correctivas
                        (inspeccion_id, item_idx, recinto, faena, item,
                         descripcion, foto_filename, fecha_limite, prioridad, estado)
                    VALUES (?,?,?,?,?,?,?,?,?,?)
                ''', (
                    inspeccion_id, i,
                    (data.get('nombre_recinto') or '').title(),
                    normalizar_faena(data.get('faena', '')),
                    item,
                    comentario or f'Ítem "{item}" marcado como NO en inspección {correlativo}',
                    foto_filename,
                    fecha_limite,
                    'Alta',
                    'Pendiente',
                ))

                # Requerimiento: incrementar si existe, crear si no
                fecha_insp_str = data.get('fecha') or datetime.now().strftime('%Y-%m-%d')
                req_faena   = normalizar_faena(data.get('faena', ''))
                req_recinto = (data.get('nombre_recinto') or '').title()
                existing = conn.execute(
                    '''SELECT id FROM requerimientos
                       WHERE item_nombre=? AND faena=? AND recinto=? AND estado='Pendiente' ''',
                    (item, req_faena, req_recinto)
                ).fetchone()
                if existing:
                    conn.execute(
                        '''UPDATE requerimientos
                           SET cantidad=cantidad+1, fecha_ultimo_reporte=?
                           WHERE id=?''',
                        (fecha_insp_str, existing['id'])
                    )
                else:
                    conn.execute(
                        '''INSERT INTO requerimientos
                               (item_nombre, cantidad, faena, recinto, fecha_ultimo_reporte, estado)
                           VALUES (?,1,?,?,?,'Pendiente')''',
                        (item, req_faena, req_recinto, fecha_insp_str)
                    )

        conn.commit()
    finally:
        conn.close()

    return redirect(url_for('exito', inspeccion_id=inspeccion_id))


@app.route('/exito/<int:inspeccion_id>')
@login_required
def exito(inspeccion_id):
    conn = get_db()
    insp = conn.execute(
        "SELECT correlativo, nombre_recinto FROM inspecciones WHERE id=?",
        (inspeccion_id,)
    ).fetchone()
    conn.close()
    return render_template('exito.html', inspeccion_id=inspeccion_id, insp=insp)


@app.route('/historial')
@login_required
def historial():
    recinto     = request.args.get('recinto', '').strip()
    fecha_desde = request.args.get('fecha_desde', '').strip()
    fecha_hasta = request.args.get('fecha_hasta', '').strip()

    if current_user.is_admin():
        query  = "SELECT * FROM inspecciones WHERE 1=1"
        params = []
    elif current_user.rol == 'supervisor':
        query  = "SELECT * FROM inspecciones WHERE faena=?"
        params = [current_user.faena]
    else:
        query  = "SELECT * FROM inspecciones WHERE usuario_id=?"
        params = [current_user.id]

    if recinto:
        query += " AND nombre_recinto LIKE ?"
        params.append(f'%{recinto}%')
    if fecha_desde:
        query += " AND fecha >= ?"
        params.append(fecha_desde)
    if fecha_hasta:
        query += " AND fecha <= ?"
        params.append(fecha_hasta)
    query += " ORDER BY created_at DESC"

    conn = get_db()
    inspecciones = conn.execute(query, params).fetchall()
    conn.close()
    return render_template('historial.html',
                           inspecciones=inspecciones,
                           recinto=recinto,
                           fecha_desde=fecha_desde,
                           fecha_hasta=fecha_hasta)


@app.route('/detalle/<int:inspeccion_id>')
@login_required
def detalle(inspeccion_id):
    conn = get_db()
    insp = conn.execute("SELECT * FROM inspecciones WHERE id=?", (inspeccion_id,)).fetchone()
    if insp is None or not _puede_ver_inspeccion(insp):
        conn.close()
        flash('Inspección no encontrada o acceso denegado.', 'danger')
        return redirect(url_for('historial'))

    items = conn.execute(
        "SELECT * FROM items_checklist WHERE inspeccion_id=? ORDER BY id",
        (inspeccion_id,)
    ).fetchall()
    fotos_rows = conn.execute(
        "SELECT item_idx, filename FROM fotos_items WHERE inspeccion_id=?",
        (inspeccion_id,)
    ).fetchall()
    conn.close()

    # Agrupar items por sección y mapear fotos por idx global
    fotos = {row['item_idx']: row['filename'] for row in fotos_rows}
    items_by_sec = {}
    global_idx = 0
    for seccion_name in SECCIONES.keys():
        items_by_sec[seccion_name] = []
        for it in items:
            if it['seccion'] == seccion_name:
                foto_url = None
                if global_idx in fotos:
                    foto_url = url_for('static', filename=f'fotos/{fotos[global_idx]}')
                items_by_sec[seccion_name].append({
                    'item':      it['item'],
                    'valor':     it['valor'],
                    'comentario':it['comentario'],
                    'foto_url':  foto_url,
                })
                global_idx += 1

    return render_template('detalle.html',
                           insp=insp,
                           items_by_sec=items_by_sec,
                           secciones=SECCIONES)


@app.route('/dashboard')
@login_required
def dashboard():
    if not current_user.is_supervisor() and not current_user.is_lectura():
        flash('Acceso restringido a Supervisores, Administradores y usuarios de Solo Lectura.', 'danger')
        return redirect(url_for('index'))

    # Filtros
    faena_filter = request.args.get('faena', '').strip()
    fecha_desde  = request.args.get('fecha_desde', '').strip()
    fecha_hasta  = request.args.get('fecha_hasta', '').strip()

    # Supervisor y lectura solo pueden ver su propia faena
    if current_user.rol in ('supervisor', 'lectura'):
        faena_filter = current_user.faena

    # Defaults: mes actual si no hay filtro de fecha
    if not fecha_desde and not fecha_hasta:
        now = datetime.now()
        fecha_desde = now.strftime('%Y-%m-01')
        fecha_hasta = now.strftime('%Y-%m-%d')

    data = _dashboard_data(faena_filter or None, fecha_desde, fecha_hasta)
    return render_template('dashboard.html',
                           data=data,
                           faenas=FAENAS,
                           faena_filter=faena_filter,
                           fecha_desde=fecha_desde,
                           fecha_hasta=fecha_hasta)


@app.route('/pdf/<int:inspeccion_id>')
@login_required
def descargar_pdf(inspeccion_id):
    conn = get_db()
    insp = conn.execute("SELECT * FROM inspecciones WHERE id=?", (inspeccion_id,)).fetchone()
    if insp is None:
        conn.close()
        return "Inspección no encontrada", 404
    if not _puede_ver_inspeccion(insp):
        conn.close()
        return "Acceso denegado", 403

    items = conn.execute(
        "SELECT * FROM items_checklist WHERE inspeccion_id=? ORDER BY id",
        (inspeccion_id,)
    ).fetchall()
    fotos_rows = conn.execute(
        "SELECT item_idx, filename FROM fotos_items WHERE inspeccion_id=?",
        (inspeccion_id,)
    ).fetchall()
    conn.close()

    fotos = {row['item_idx']: row['filename'] for row in fotos_rows}
    buf = generate_pdf(insp, items, fotos)
    filename = f'INCO-INI-VH-001_{insp["correlativo"]}.pdf'
    return send_file(buf, as_attachment=True, download_name=filename, mimetype='application/pdf')


# ─── Acciones Correctivas ────────────────────────────────────────────────────

ESTADOS_AC   = ['Pendiente', 'En proceso', 'Completada']
PRIORIDADES  = ['Alta', 'Media', 'Baja']


def _acciones_pendientes_usuario():
    """Cuenta acciones pendientes/en-proceso asignadas al usuario actual."""
    if not current_user.is_authenticated:
        return 0
    conn = get_db()
    n = conn.execute(
        "SELECT COUNT(*) FROM acciones_correctivas "
        "WHERE responsable_id=? AND estado != 'Completada'",
        (current_user.id,)
    ).fetchone()[0]
    conn.close()
    return n


# Inyectar contador y utilidades en todos los templates
@app.context_processor
def inject_globals():
    count = _acciones_pendientes_usuario() if current_user.is_authenticated else 0
    return {'acciones_pendientes_badge': count, 'normalizar_faena': normalizar_faena}


@app.route('/acciones')
@login_required
def acciones():
    estado_f    = request.args.get('estado',   '').strip()
    prioridad_f = request.args.get('prioridad','').strip()
    faena_f     = request.args.get('faena',    '').strip()
    resp_f      = request.args.get('responsable', '').strip()

    q  = '''SELECT ac.*, u.nombre_completo AS responsable_nombre
            FROM acciones_correctivas ac
            LEFT JOIN usuarios u ON ac.responsable_id = u.id
            WHERE 1=1'''
    p  = []

    # Filtro por rol
    if current_user.rol == 'inspector':
        q += ' AND ac.inspeccion_id IN (SELECT id FROM inspecciones WHERE usuario_id=?)'
        p.append(current_user.id)
    elif current_user.rol == 'supervisor':
        q += ' AND ac.faena=?'
        p.append(current_user.faena)

    if estado_f:
        q += ' AND ac.estado=?'; p.append(estado_f)
    if prioridad_f:
        q += ' AND ac.prioridad=?'; p.append(prioridad_f)
    if faena_f:
        q += ' AND ac.faena=?'; p.append(faena_f)
    if resp_f:
        q += ' AND ac.responsable_id=?'; p.append(resp_f)

    q += ' ORDER BY CASE ac.estado WHEN "Pendiente" THEN 1 WHEN "En proceso" THEN 2 ELSE 3 END,'
    q += ' CASE ac.prioridad WHEN "Alta" THEN 1 WHEN "Media" THEN 2 ELSE 3 END,'
    q += ' ac.fecha_limite ASC'

    conn = get_db()
    acciones_list = conn.execute(q, p).fetchall()
    usuarios_list = conn.execute(
        "SELECT id, nombre_completo, rol FROM usuarios WHERE activo=1 ORDER BY nombre_completo"
    ).fetchall()
    conn.close()

    today = datetime.now().strftime('%Y-%m-%d')
    return render_template('acciones.html',
                           acciones=acciones_list,
                           usuarios=usuarios_list,
                           estados=ESTADOS_AC,
                           prioridades=PRIORIDADES,
                           faenas=FAENAS,
                           estado_f=estado_f, prioridad_f=prioridad_f,
                           faena_f=faena_f, resp_f=resp_f,
                           today=today)


@app.route('/acciones/<int:accion_id>')
@login_required
def accion_detalle(accion_id):
    conn = get_db()
    ac = conn.execute(
        '''SELECT ac.*, u.nombre_completo AS responsable_nombre,
                  i.correlativo, i.nombre_recinto
           FROM acciones_correctivas ac
           LEFT JOIN usuarios u ON ac.responsable_id = u.id
           LEFT JOIN inspecciones i ON ac.inspeccion_id = i.id
           WHERE ac.id=?''', (accion_id,)
    ).fetchone()
    if ac is None:
        conn.close()
        flash('Acción no encontrada.', 'danger')
        return redirect(url_for('acciones'))

    seguimientos = conn.execute(
        '''SELECT s.*, u.nombre_completo AS autor
           FROM seguimiento_acciones s
           JOIN usuarios u ON s.usuario_id = u.id
           WHERE s.accion_id=? ORDER BY s.created_at ASC''',
        (accion_id,)
    ).fetchall()
    usuarios_list = conn.execute(
        "SELECT id, nombre_completo FROM usuarios WHERE activo=1 ORDER BY nombre_completo"
    ).fetchall()
    conn.close()

    today = datetime.now().strftime('%Y-%m-%d')
    return render_template('accion_detalle.html',
                           ac=ac,
                           seguimientos=seguimientos,
                           usuarios=usuarios_list,
                           estados=ESTADOS_AC,
                           prioridades=PRIORIDADES,
                           today=today)


@app.route('/acciones/<int:accion_id>/actualizar', methods=['POST'])
@login_required
@escritura_required
def accion_actualizar(accion_id):
    conn = get_db()
    ac = conn.execute('SELECT * FROM acciones_correctivas WHERE id=?', (accion_id,)).fetchone()
    if ac is None:
        conn.close()
        flash('Acción no encontrada.', 'danger')
        return redirect(url_for('acciones'))

    f             = request.form
    nuevo_estado  = f.get('estado',         ac['estado'])
    nueva_prio    = f.get('prioridad',       ac['prioridad'])
    nuevo_resp    = f.get('responsable_id',  '') or None
    nueva_fechal  = f.get('fecha_limite',    ac['fecha_limite'] or '')
    observacion   = f.get('observacion',     '').strip()

    fecha_cierre = ac['fecha_cierre']
    if nuevo_estado == 'Completada' and ac['estado'] != 'Completada':
        fecha_cierre = datetime.now().strftime('%Y-%m-%d %H:%M:%S')
    elif nuevo_estado != 'Completada':
        fecha_cierre = None

    conn.execute('''
        UPDATE acciones_correctivas
        SET estado=?, prioridad=?, responsable_id=?, fecha_limite=?,
            fecha_cierre=?
        WHERE id=?
    ''', (nuevo_estado, nueva_prio, nuevo_resp, nueva_fechal or None,
          fecha_cierre, accion_id))

    if observacion:
        conn.execute('''
            INSERT INTO seguimiento_acciones (accion_id, usuario_id, observacion, estado_nuevo)
            VALUES (?,?,?,?)
        ''', (accion_id, current_user.id, observacion, nuevo_estado))

    conn.commit()
    conn.close()
    flash('Acción actualizada correctamente.', 'success')
    return redirect(url_for('accion_detalle', accion_id=accion_id))


@app.route('/acciones/<int:accion_id>/asignar', methods=['POST'])
@login_required
@escritura_required
def accion_asignar(accion_id):
    """Asignación desde detalle: responsable, prioridad y fecha límite."""
    responsable_id = request.form.get('responsable_id') or None
    prioridad      = request.form.get('prioridad', 'Alta')
    fecha_limite   = request.form.get('fecha_limite') or None
    conn = get_db()
    conn.execute(
        'UPDATE acciones_correctivas SET responsable_id=?, prioridad=?, fecha_limite=? WHERE id=?',
        (responsable_id, prioridad, fecha_limite, accion_id)
    )
    conn.commit()
    conn.close()
    flash('Asignación actualizada.', 'success')
    return redirect(url_for('accion_detalle', accion_id=accion_id))


# ─── Módulo 5: Requerimientos de Materiales ──────────────────────────────────

@app.route('/requerimientos')
@login_required
def requerimientos():
    if not current_user.is_supervisor() and not current_user.is_lectura():
        flash('Acceso restringido a Supervisores, Administradores y usuarios de Solo Lectura.', 'danger')
        return redirect(url_for('index'))

    faena_f  = request.args.get('faena',  '').strip()
    estado_f = request.args.get('estado', '').strip()

    q = '''SELECT * FROM requerimientos WHERE 1=1'''
    p = []
    if faena_f:
        q += ' AND faena=?'
        p.append(faena_f)
    if current_user.rol in ('supervisor', 'lectura'):
        q += ' AND faena=?'
        p.append(current_user.faena)
    if estado_f:
        q += ' AND estado=?'
        p.append(estado_f)
    q += ' ORDER BY cantidad DESC, fecha_ultimo_reporte DESC'

    conn = get_db()
    reqs = conn.execute(q, p).fetchall()

    # KPIs
    kpi_q  = 'SELECT * FROM requerimientos WHERE estado=\'Pendiente\''
    kpi_p  = []
    if current_user.rol in ('supervisor', 'lectura'):
        kpi_q += ' AND faena=?'
        kpi_p.append(current_user.faena)
    pendientes_list = conn.execute(kpi_q + ' ORDER BY cantidad DESC', kpi_p).fetchall()
    total_pendientes = len(pendientes_list)
    mas_recurrente   = pendientes_list[0]['item_nombre'] if pendientes_list else '—'
    max_cantidad     = pendientes_list[0]['cantidad']    if pendientes_list else 0

    faenas_list = [r[0] for r in conn.execute(
        'SELECT DISTINCT faena FROM requerimientos WHERE faena IS NOT NULL ORDER BY faena'
    ).fetchall()]
    conn.close()

    return render_template('requerimientos.html',
                           reqs=reqs,
                           faenas=faenas_list,
                           faena_f=faena_f,
                           estado_f=estado_f,
                           total_pendientes=total_pendientes,
                           mas_recurrente=mas_recurrente,
                           max_cantidad=max_cantidad)


@app.route('/requerimientos/exportar')
@login_required
def req_exportar():
    if not current_user.is_supervisor() and not current_user.is_lectura():
        flash('Acceso restringido.', 'danger')
        return redirect(url_for('index'))

    import openpyxl
    from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
    from openpyxl.utils import get_column_letter

    faena_f  = request.args.get('faena',  '').strip()
    estado_f = request.args.get('estado', '').strip()

    q = 'SELECT * FROM requerimientos WHERE 1=1'
    p = []
    if faena_f:
        q += ' AND faena=?'
        p.append(faena_f)
    if current_user.rol in ('supervisor', 'lectura'):
        q += ' AND faena=?'
        p.append(current_user.faena)
    if estado_f:
        q += ' AND estado=?'
        p.append(estado_f)
    q += ' ORDER BY cantidad DESC, fecha_ultimo_reporte DESC'

    conn = get_db()
    reqs = conn.execute(q, p).fetchall()
    conn.close()

    # ── Estilos ─────────────────────────────────────────────────────────────
    AZUL_HDR    = PatternFill('solid', fgColor='1A3A5C')
    ROJO_CLARO  = PatternFill('solid', fgColor='FFCDD2')
    AMAR_CLARO  = PatternFill('solid', fgColor='FFF9C4')
    VERDE_CLARO = PatternFill('solid', fgColor='C8E6C9')
    BLANCO      = PatternFill('solid', fgColor='FFFFFF')
    thin  = Side(style='thin', color='BDBDBD')
    borde = Border(left=thin, right=thin, top=thin, bottom=thin)

    # ── Libro ────────────────────────────────────────────────────────────────
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = 'Requerimientos INCOVALL'

    # Fila 1: Título fusionado
    ws.merge_cells('A1:G1')
    c = ws['A1']
    c.value     = 'Solicitud de Materiales — INCOVALL'
    c.font      = Font(bold=True, size=14, color='1A3A5C')
    c.alignment = Alignment(horizontal='center', vertical='center')
    ws.row_dimensions[1].height = 28

    # Fila 2: metadatos
    fecha_gen = datetime.now().strftime('%d-%m-%Y')
    ws['A2'] = f'Fecha de generación: {fecha_gen}'
    ws['A2'].font = Font(italic=True, size=10, color='555555')
    ws['D2'] = f'Generado por: {current_user.nombre_completo}'
    ws['D2'].font = Font(italic=True, size=10, color='555555')

    # Fila 3: vacía
    ws.row_dimensions[3].height = 8

    # Fila 4: encabezados
    headers = ['N°', 'Ítem defectuoso', 'Faena', 'Recinto', 'Cantidad', 'Estado', 'Último reporte']
    for col, hdr in enumerate(headers, 1):
        cell = ws.cell(row=4, column=col, value=hdr)
        cell.font      = Font(bold=True, color='FFFFFF', size=11)
        cell.fill      = AZUL_HDR
        cell.alignment = Alignment(horizontal='center', vertical='center')
        cell.border    = borde
    ws.row_dimensions[4].height = 20

    # Filas de datos
    for idx, r in enumerate(reqs, 1):
        row_num  = idx + 4
        cantidad = r['cantidad']
        estado   = r['estado']
        if estado == 'Gestionado':
            fill = VERDE_CLARO
        elif cantidad >= 3:
            fill = ROJO_CLARO
        elif cantidad == 2:
            fill = AMAR_CLARO
        else:
            fill = BLANCO
        vals = [idx, r['item_nombre'], r['faena'] or '', r['recinto'] or '',
                cantidad, estado, r['fecha_ultimo_reporte'] or '']
        for col, val in enumerate(vals, 1):
            cell = ws.cell(row=row_num, column=col, value=val)
            cell.fill      = fill
            cell.border    = borde
            cell.alignment = Alignment(vertical='center',
                                       horizontal='center' if col in (1, 5) else 'left')

    # Última fila: total pendientes
    total_pend = sum(1 for r in reqs if r['estado'] == 'Pendiente')
    last_row = len(reqs) + 5
    ws.merge_cells(f'A{last_row}:D{last_row}')
    c = ws.cell(row=last_row, column=1, value=f'Total ítems pendientes: {total_pend}')
    c.font      = Font(bold=True, size=11)
    c.alignment = Alignment(horizontal='left', vertical='center')
    ws.row_dimensions[last_row].height = 18

    # Anchos de columna
    col_widths = [5, 32, 20, 26, 10, 14, 18]
    for col, width in enumerate(col_widths, 1):
        ws.column_dimensions[get_column_letter(col)].width = width

    # Congelar encabezados
    ws.freeze_panes = 'A5'

    # ── Respuesta ────────────────────────────────────────────────────────────
    output = io.BytesIO()
    wb.save(output)
    output.seek(0)
    fname_suffix = f'_{faena_f}' if faena_f else ''
    fname = f'Requerimientos_INCOVALL{fname_suffix}_{datetime.now().strftime("%Y%m%d")}.xlsx'
    return send_file(
        output,
        mimetype='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet',
        as_attachment=True,
        download_name=fname
    )


@app.route('/requerimientos/<int:req_id>/gestionar', methods=['POST'])
@login_required
@escritura_required
def req_gestionar(req_id):
    if not current_user.is_supervisor():
        flash('Acceso restringido.', 'danger')
        return redirect(url_for('index'))
    conn = get_db()
    conn.execute("UPDATE requerimientos SET estado='Gestionado' WHERE id=?", (req_id,))
    conn.commit()
    conn.close()
    flash('Requerimiento marcado como Gestionado.', 'success')
    return redirect(request.referrer or url_for('requerimientos'))


# ─── Admin: Gestión de Usuarios ──────────────────────────────────────────────

@app.route('/admin/usuarios')
@login_required
@admin_required
def admin_usuarios():
    conn = get_db()
    usuarios = conn.execute("SELECT * FROM usuarios ORDER BY rol, nombre_completo").fetchall()
    conn.close()
    return render_template('admin_usuarios.html', usuarios=usuarios)


@app.route('/admin/usuarios/nuevo', methods=['GET', 'POST'])
@login_required
@admin_required
def admin_usuario_nuevo():
    if request.method == 'POST':
        return _guardar_usuario(None)
    return render_template('admin_usuario_form.html',
                           accion='Nuevo', usuario=None,
                           roles=ROLES, faenas=FAENAS)


@app.route('/admin/usuarios/<int:uid>/editar', methods=['GET', 'POST'])
@login_required
@admin_required
def admin_usuario_editar(uid):
    conn = get_db()
    usuario = conn.execute("SELECT * FROM usuarios WHERE id=?", (uid,)).fetchone()
    conn.close()
    if usuario is None:
        flash('Usuario no encontrado.', 'danger')
        return redirect(url_for('admin_usuarios'))
    if request.method == 'POST':
        return _guardar_usuario(uid)
    return render_template('admin_usuario_form.html',
                           accion='Editar', usuario=usuario,
                           roles=ROLES, faenas=FAENAS)


@app.route('/admin/usuarios/<int:uid>/toggle', methods=['POST'])
@login_required
@admin_required
def admin_usuario_toggle(uid):
    if uid == current_user.id:
        flash('No puedes desactivar tu propia cuenta.', 'warning')
        return redirect(url_for('admin_usuarios'))
    conn = get_db()
    conn.execute("UPDATE usuarios SET activo = 1 - activo WHERE id=?", (uid,))
    conn.commit()
    conn.close()
    flash('Estado del usuario actualizado.', 'success')
    return redirect(url_for('admin_usuarios'))


def _guardar_usuario(uid):
    f = request.form
    username        = f.get('username', '').strip()
    nombre_completo = f.get('nombre_completo', '').strip()
    email           = f.get('email', '').strip()
    rol             = f.get('rol', 'inspector')
    faena           = f.get('faena', '').strip() or None
    activo          = 1 if f.get('activo') else 0
    password        = f.get('password', '').strip()

    if not username or not nombre_completo:
        flash('Usuario y nombre son obligatorios.', 'danger')
        return redirect(request.url)

    conn = get_db()
    try:
        if uid is None:
            if not password:
                flash('La contraseña es obligatoria para un usuario nuevo.', 'danger')
                return redirect(request.url)
            conn.execute('''
                INSERT INTO usuarios (username, nombre_completo, email, password_hash, rol, faena, activo)
                VALUES (?,?,?,?,?,?,?)
            ''', (username, nombre_completo, email,
                  generate_password_hash(password), rol, faena, activo))
            flash(f'Usuario "{username}" creado correctamente.', 'success')
        else:
            if password:
                conn.execute('''
                    UPDATE usuarios SET username=?, nombre_completo=?, email=?,
                        password_hash=?, rol=?, faena=?, activo=?
                    WHERE id=?
                ''', (username, nombre_completo, email,
                      generate_password_hash(password), rol, faena, activo, uid))
            else:
                conn.execute('''
                    UPDATE usuarios SET username=?, nombre_completo=?, email=?,
                        rol=?, faena=?, activo=?
                    WHERE id=?
                ''', (username, nombre_completo, email, rol, faena, activo, uid))
            flash(f'Usuario "{username}" actualizado correctamente.', 'success')
        conn.commit()
    except sqlite3.IntegrityError:
        flash(f'El nombre de usuario "{username}" ya está en uso.', 'danger')
        return redirect(request.url)
    finally:
        conn.close()

    return redirect(url_for('admin_usuarios'))


# ─── PDF Generation ──────────────────────────────────────────────────────────

C_AZUL       = colors.HexColor('#0d47a1')
C_AZUL_SEC   = colors.HexColor('#1565C0')
C_AZUL_LIGHT = colors.HexColor('#E3F2FD')
C_VERDE      = colors.HexColor('#388E3C')
C_ROJO       = colors.HexColor('#C62828')
C_GRIS       = colors.HexColor('#757575')
C_VERDE_BG   = colors.HexColor('#C8E6C9')
C_ROJO_BG    = colors.HexColor('#FFCDD2')
C_GRIS_BG    = colors.HexColor('#E0E0E0')
C_LINEA      = colors.HexColor('#BDBDBD')
C_FILA_ALT   = colors.HexColor('#F5F5F5')
C_FOTO_BG    = colors.HexColor('#FFF3E0')


def pdf_logo(filename, fallback_text, max_w=4.0, max_h=2.2):
    path = os.path.join(STATIC_DIR, filename)
    if os.path.exists(path):
        try:
            with PILImage.open(path) as pil_img:
                pil_img.load()
                nat_w, nat_h = pil_img.size
            ratio = min(max_w * cm / nat_w, max_h * cm / nat_h)
            img = Image(path, width=nat_w * ratio, height=nat_h * ratio)
            img.hAlign = 'CENTER'
            return img
        except Exception:
            pass
    w, h = max_w * cm, max_h * cm
    d = Drawing(w, h)
    d.add(Rect(0, 0, w, h, fillColor=C_AZUL, strokeColor=None))
    d.add(RLString(8, h * 0.52, fallback_text,
                   fontName='Helvetica-Bold', fontSize=15, fillColor=colors.white))
    return d


def _pdf_foto(filename, max_w_cm=7.0, max_h_cm=5.0):
    """Devuelve un Image de reportlab para una foto, o None si no existe/falla."""
    path = os.path.join(FOTOS_DIR, filename)
    if not os.path.exists(path):
        return None
    try:
        with PILImage.open(path) as pil_img:
            pil_img.load()
            nat_w, nat_h = pil_img.size
        ratio = min(max_w_cm * cm / nat_w, max_h_cm * cm / nat_h)
        img = Image(path, width=nat_w * ratio, height=nat_h * ratio)
        img.hAlign = 'LEFT'
        return img
    except Exception:
        return None


def generate_pdf(insp, items, fotos=None):
    if fotos is None:
        fotos = {}

    buf = BytesIO()
    doc = SimpleDocTemplate(buf, pagesize=A4,
                            rightMargin=1.5 * cm, leftMargin=1.5 * cm,
                            topMargin=1.5 * cm, bottomMargin=1.5 * cm)
    styles = getSampleStyleSheet()
    W = 18 * cm

    def style(name, **kw):
        return ParagraphStyle(name, parent=styles['Normal'], **kw)

    s_white_bold = style('wb',  fontSize=9,  fontName='Helvetica-Bold',
                         textColor=colors.white, alignment=TA_CENTER)
    s_label      = style('lbl', fontSize=7.5, textColor=colors.HexColor('#555'),
                         fontName='Helvetica')
    s_value      = style('val', fontSize=9,  fontName='Helvetica-Bold')
    s_sec_header = style('sh',  fontSize=8.5, fontName='Helvetica-Bold',
                         textColor=colors.white)
    s_item       = style('it',  fontSize=8.5, fontName='Helvetica')
    s_comment    = style('cm',  fontSize=7.5, fontName='Helvetica',
                         textColor=colors.HexColor('#444'))
    s_mark_si    = style('msi', fontSize=13, fontName='Helvetica-Bold',
                         textColor=C_VERDE, alignment=TA_CENTER)
    s_mark_no    = style('mno', fontSize=13, fontName='Helvetica-Bold',
                         textColor=C_ROJO, alignment=TA_CENTER)
    s_mark_na    = style('mna', fontSize=13, fontName='Helvetica-Bold',
                         textColor=C_GRIS, alignment=TA_CENTER)
    s_mark_empty = style('me',  fontSize=13, alignment=TA_CENTER)
    s_footer     = style('ft',  fontSize=7,  textColor=colors.HexColor('#999'),
                         alignment=TA_CENTER)
    s_obs_body   = style('ob',  fontSize=9,  fontName='Helvetica', leading=13)
    s_obs_label  = style('ol',  fontSize=9,  fontName='Helvetica-Bold',
                         textColor=colors.white)
    s_foto_lbl   = style('fl',  fontSize=7,  fontName='Helvetica-Bold',
                         textColor=colors.HexColor('#E65100'))

    story = []

    # ── Cabecera ────────────────────────────────────────────────────────────
    s_title_main = style('tm', fontSize=13, fontName='Helvetica-Bold',
                         textColor=colors.white, alignment=TA_CENTER)
    s_title_sub  = style('ts', fontSize=8.5, fontName='Helvetica',
                         textColor=colors.HexColor('#BBDEFB'), alignment=TA_CENTER)
    s_title_info = style('ti', fontSize=8, fontName='Helvetica',
                         textColor=colors.HexColor('#90CAF9'), alignment=TA_CENTER)

    logo_incovall = pdf_logo('logo_incovall.png', 'INCOVALL', max_w=4.0, max_h=2.2)
    logo_cmp      = pdf_logo('logo_cmp.png',      'CMP',      max_w=4.0, max_h=2.2)

    title_block = [
        Paragraph('Inspección Preventiva INI', s_title_main),
        Spacer(1, 4),
        Paragraph('INCO-INI-VH-001', s_title_sub),
        Spacer(1, 6),
        Paragraph(
            f'Correlativo: <b>{insp["correlativo"]}</b> &nbsp;|&nbsp; '
            f'Fecha: <b>{insp["fecha"] or ""}</b>',
            s_title_info
        ),
    ]

    header_data = [[logo_incovall, title_block, logo_cmp]]
    header_t = Table(header_data, colWidths=[4.0 * cm, 10.0 * cm, 4.0 * cm])
    header_t.setStyle(TableStyle([
        ('BACKGROUND',    (0, 0), (-1, -1), C_AZUL),
        ('VALIGN',        (0, 0), (-1, -1), 'MIDDLE'),
        ('ALIGN',         (0, 0), (-1, -1), 'CENTER'),
        ('TOPPADDING',    (0, 0), (-1, -1), 10),
        ('BOTTOMPADDING', (0, 0), (-1, -1), 10),
        ('LEFTPADDING',   (0, 0), (-1, -1), 6),
        ('RIGHTPADDING',  (0, 0), (-1, -1), 6),
    ]))
    story.append(header_t)
    story.append(Spacer(1, 0.35 * cm))

    # ── Ficha de datos ───────────────────────────────────────────────────────
    def field(label, value):
        return [Paragraph(label, s_label), Paragraph(str(value or ''), s_value)]

    info_rows = [
        field('VALLE', insp['valle']) + field('FAENA', insp['faena']) + field('CLIENTE', insp['cliente']),
        field('RECINTO', insp['nombre_recinto']) + field('INSPECTOR', insp['nombre_inspector']) + ['', ''],
    ]
    col_w = [2.2 * cm, 3.8 * cm, 2 * cm, 3.8 * cm, 2 * cm, 4.2 * cm]
    info_t = Table(info_rows, colWidths=col_w)
    info_t.setStyle(TableStyle([
        ('BOX',           (0, 0), (-1, -1), 0.5, C_LINEA),
        ('INNERGRID',     (0, 0), (-1, -1), 0.3, C_LINEA),
        ('BACKGROUND',    (0, 0), (-1, 0),  C_AZUL_LIGHT),
        ('BACKGROUND',    (0, 1), (-1, 1),  colors.white),
        ('VALIGN',        (0, 0), (-1, -1), 'MIDDLE'),
        ('TOPPADDING',    (0, 0), (-1, -1), 4),
        ('BOTTOMPADDING', (0, 0), (-1, -1), 4),
        ('LEFTPADDING',   (0, 0), (-1, -1), 5),
        ('SPAN',          (3, 1), (5, 1)),
    ]))
    story.append(info_t)
    story.append(Spacer(1, 0.35 * cm))

    # ── Checklist ────────────────────────────────────────────────────────────
    items_by_sec = {}
    for row in items:
        items_by_sec.setdefault(row['seccion'], []).append(row)

    COL_ITEM = 5.8 * cm
    COL_CHK  = 1.6 * cm
    COL_CMT  = W - COL_ITEM - 3 * COL_CHK

    tbl_data = [[
        Paragraph('<b>ÍTEM</b>',        s_white_bold),
        Paragraph('<b>SI</b>',          s_white_bold),
        Paragraph('<b>NO</b>',          s_white_bold),
        Paragraph('<b>N/A</b>',         s_white_bold),
        Paragraph('<b>COMENTARIOS</b>', s_white_bold),
    ]]
    tbl_styles = [
        ('BACKGROUND',    (0, 0), (-1, 0),  C_AZUL),
        ('TEXTCOLOR',     (0, 0), (-1, 0),  colors.white),
        ('FONTSIZE',      (0, 0), (-1, -1), 8.5),
        ('BOX',           (0, 0), (-1, -1), 0.5, C_LINEA),
        ('INNERGRID',     (0, 0), (-1, -1), 0.3, C_LINEA),
        ('VALIGN',        (0, 0), (-1, -1), 'MIDDLE'),
        ('TOPPADDING',    (0, 0), (-1, -1), 4),
        ('BOTTOMPADDING', (0, 0), (-1, -1), 4),
        ('LEFTPADDING',   (0, 0), (-1, -1), 5),
    ]

    row_idx    = 1
    global_idx = 0

    for seccion_name in SECCIONES.keys():
        tbl_data.append([Paragraph(seccion_name, s_sec_header), '', '', '', ''])
        tbl_styles += [
            ('BACKGROUND', (0, row_idx), (-1, row_idx), C_AZUL_SEC),
            ('SPAN',       (0, row_idx), (-1, row_idx)),
        ]
        row_idx += 1

        for it in items_by_sec.get(seccion_name, []):
            valor = it['valor'] or ''
            si  = Paragraph('✓', s_mark_si)    if valor == 'SI'  else Paragraph('', s_mark_empty)
            no  = Paragraph('✓', s_mark_no)    if valor == 'NO'  else Paragraph('', s_mark_empty)
            na  = Paragraph('✓', s_mark_na)    if valor == 'N/A' else Paragraph('', s_mark_empty)

            # Celda de comentario: texto + foto embebida si valor == NO
            comment_content = [Paragraph(it['comentario'] or '', s_comment)]
            has_foto = False

            if valor == 'NO' and global_idx in fotos:
                foto_img = _pdf_foto(fotos[global_idx], max_w_cm=COL_CMT / cm - 0.3, max_h_cm=4.5)
                if foto_img is not None:
                    comment_content.append(Spacer(1, 3))
                    comment_content.append(Paragraph('Evidencia:', s_foto_lbl))
                    comment_content.append(Spacer(1, 2))
                    comment_content.append(foto_img)
                    has_foto = True

            tbl_data.append([
                Paragraph(it['item'], s_item),
                si, no, na,
                comment_content,
            ])

            if row_idx % 2 == 0 and not has_foto:
                tbl_styles.append(('BACKGROUND', (0, row_idx), (-1, row_idx), C_FILA_ALT))
            if valor == 'SI':
                tbl_styles.append(('BACKGROUND', (1, row_idx), (1, row_idx), C_VERDE_BG))
            elif valor == 'NO':
                tbl_styles.append(('BACKGROUND', (2, row_idx), (2, row_idx), C_ROJO_BG))
                if has_foto:
                    tbl_styles.append(('BACKGROUND', (0, row_idx), (-1, row_idx), C_FOTO_BG))
            elif valor == 'N/A':
                tbl_styles.append(('BACKGROUND', (3, row_idx), (3, row_idx), C_GRIS_BG))

            tbl_styles.append(('ALIGN',  (1, row_idx), (3, row_idx), 'CENTER'))
            tbl_styles.append(('VALIGN', (0, row_idx), (-1, row_idx), 'TOP'))
            row_idx    += 1
            global_idx += 1

    checklist_t = Table(tbl_data, colWidths=[COL_ITEM, COL_CHK, COL_CHK, COL_CHK, COL_CMT])
    checklist_t.setStyle(TableStyle(tbl_styles))
    story.append(checklist_t)
    story.append(Spacer(1, 0.4 * cm))

    # ── Observaciones ────────────────────────────────────────────────────────
    obs_t = Table([
        [Paragraph('OBSERVACIONES GENERALES:', s_obs_label)],
        [Paragraph(insp['observaciones'] or ' ', s_obs_body)],
    ], colWidths=[W])
    obs_t.setStyle(TableStyle([
        ('BACKGROUND',    (0, 0), (-1, 0), C_AZUL),
        ('BACKGROUND',    (0, 1), (-1, 1), colors.white),
        ('BOX',           (0, 0), (-1, -1), 0.5, C_LINEA),
        ('LEFTPADDING',   (0, 0), (-1, -1), 7),
        ('RIGHTPADDING',  (0, 0), (-1, -1), 7),
        ('TOPPADDING',    (0, 0), (-1, -1), 5),
        ('BOTTOMPADDING', (0, 0), (-1, -1), 5),
        ('MINROWHEIGHT',  (0, 1), (-1, 1), 1.8 * cm),
        ('VALIGN',        (0, 1), (-1, 1), 'TOP'),
    ]))
    story.append(obs_t)
    story.append(Spacer(1, 0.4 * cm))

    # ── Supervisor / Firma ───────────────────────────────────────────────────
    s_sup_lbl = style('sl', fontSize=9, fontName='Helvetica-Bold',
                      textColor=colors.HexColor('#333'))
    s_sup_val = style('sv', fontSize=10, fontName='Helvetica-Bold', textColor=C_AZUL)
    s_linea   = style('ln', fontSize=8, textColor=colors.HexColor('#888'),
                      alignment=TA_CENTER)

    firma_t = Table([
        [Paragraph('SUPERVISOR RESPONSABLE:', s_sup_lbl),
         Paragraph(insp['supervisor'] or '', s_sup_val)],
        ['', Paragraph('___________________________________', s_linea)],
        ['', Paragraph('Firma', s_linea)],
    ], colWidths=[6 * cm, 12 * cm])
    firma_t.setStyle(TableStyle([
        ('BOX',           (0, 0), (-1, -1), 0.5, C_LINEA),
        ('VALIGN',        (0, 0), (-1, -1), 'MIDDLE'),
        ('TOPPADDING',    (0, 0), (-1, -1), 5),
        ('BOTTOMPADDING', (0, 0), (-1, -1), 5),
        ('LEFTPADDING',   (0, 0), (-1, -1), 7),
        ('ALIGN',         (1, 1), (1, 2),   'CENTER'),
    ]))
    story.append(firma_t)
    story.append(Spacer(1, 0.3 * cm))

    story.append(Paragraph(
        f'Generado el {datetime.now().strftime("%d/%m/%Y %H:%M")} &nbsp;|&nbsp; '
        f'INCOVALL – Inspección Preventiva INI &nbsp;|&nbsp; INCO-INI-VH-001',
        s_footer
    ))

    doc.build(story)
    buf.seek(0)
    return buf


# ─── Startup ─────────────────────────────────────────────────────────────────
# Se llama aquí para que Gunicorn (que importa el módulo directamente) también
# inicialice la base de datos al arrancar.
try:
    init_db()
except Exception as _e:
    import logging
    logging.getLogger(__name__).error('init_db() falló al arrancar: %s', _e, exc_info=True)

# ─── Main ────────────────────────────────────────────────────────────────────

if __name__ == '__main__':
    import webbrowser, threading
    def open_browser():
        import time; time.sleep(1.2)
        webbrowser.open('http://localhost:5000')
    threading.Thread(target=open_browser, daemon=True).start()
    print("\n  INCOVALL - Inspección Preventiva INI")
    print("  Servidor en http://localhost:5000")
    print("  Presione Ctrl+C para detener.\n")
    app.run(debug=False, host='0.0.0.0', port=5000)
